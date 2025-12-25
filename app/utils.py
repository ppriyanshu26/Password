from pathlib import Path
from openpyxl import Workbook
from classes import Crypto
from credentials import load_vault
from urllib.parse import urlparse, parse_qs
import re
import cv2

def extract_totp_from_qr(image_path):
    try:
        payloads = []
        img_cv = cv2.imread(str(image_path))
        if img_cv is not None:
            detector = cv2.QRCodeDetector()
            try:
                retval, decoded_infos, points, _ = detector.detectAndDecodeMulti(img_cv)
                if retval and decoded_infos:
                    for d in decoded_infos:
                        if d:
                            payloads.append(d.strip())
            except Exception:
                data, points, _ = detector.detectAndDecode(img_cv)
                if data:
                    payloads.append(data.strip())

        if not payloads:
            return None

        for data in payloads:
            if not data:
                continue
            if data.lower().startswith("otpauth://"):
                qs = urlparse(data).query
                params = parse_qs(qs)
                secret = params.get("secret", [None])[0]
                if secret:
                    return secret
            m = re.search(r"secret=([A-Z2-7]+=*)", data, re.IGNORECASE)
            if m:
                return m.group(1)
            candidate = re.sub(r"[^A-Z2-7=]", "", data.upper())
            if len(candidate) >= 16:
                return candidate
            if len(data) >= 8:
                return data
    except Exception:
        return None

def export_credentials_to_excel(master_key):
    desktop_path = Path.home() / "Desktop"
    file_path = desktop_path / "credentials.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws['A1'], ws['B1'], ws['C1'], ws['D1'] = "Platform", "Username", "Password", "MFA"
    
    crypto = Crypto(master_key)
    vault = load_vault()
    row = 2
    
    for platform, accounts in vault.items():
        for account in accounts:
            ws[f'A{row}'] = platform
            ws[f'B{row}'] = account['username']
            ws[f'C{row}'] = crypto.decrypt_aes(account['password'])
            ws[f'D{row}'] = crypto.decrypt_aes(account['mfa']) if account.get('mfa') else ""
            row += 1
    
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 25
    
    wb.save(file_path)
    return {"success": True, "message": "Exported successfully", "file_path": str(file_path)}
