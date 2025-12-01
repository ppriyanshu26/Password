import json
import pyotp
import time
import os
from openpyxl import Workbook
from openpyxl.styles import Font
import keyring
import hashlib
import getpass

APP_FOLDER = os.path.join(os.getenv('APPDATA'), "Password Filler")
os.makedirs(APP_FOLDER, exist_ok=True)
FILE = os.path.join(APP_FOLDER, "credentials.json")

SERVICE_NAME = "Password Filler"

def check_key(password, service=SERVICE_NAME):
    username = getpass.getuser()
    stored = keyring.get_password(service, username)
    if not stored:
        username = getpass.getuser()
        hashed = hashlib.sha256(password.encode()).hexdigest()
        keyring.set_password(service, username, hashed)
        return True
    return hashlib.sha256(password.encode()).hexdigest() == stored

if not os.path.exists(FILE):
    with open(FILE, "w") as f:
        json.dump({}, f, indent=4)

def generate_totp(secret):
    return pyotp.TOTP(secret).now()
    
def load_creds():
    with open(FILE, 'r') as f:
        return json.load(f)

def save_creds(creds):
    sorted_creds = dict(sorted(creds.items(), key=lambda x: x[0].lower()))
    with open(FILE, 'w') as f:
        json.dump(sorted_creds, f, indent=4)

def add_credential(app, username, password, secretco, crypto):
    creds = load_creds()
    new_cred = {"username": username, "password": crypto.encrypt_aes256(password)}
    if secretco:
        new_cred["secretco"] = crypto.encrypt_aes256(secretco)
    app = app.strip().lower()
    if app in creds:
        creds[app].append(new_cred)
    else:
        creds[app] = [new_cred]
    save_creds(creds)
    return True, f"Credential added for '{app.capitalize()}'!"

def delete_credential(app, username):
    creds = load_creds()
    original_count = len(creds[app])
    creds[app] = [c for c in creds[app] if c['username'] != username]
    if len(creds[app]) == original_count:
        return False, "Credential not found!"
    if not creds[app]:
        del creds[app]
    save_creds(creds)
    return True, f"Credential deleted for '{app.capitalize()}'!"

def download(crypto):
    try:
        creds = load_creds()
        wb = Workbook()
        ws = wb.active
        ws.title = "Credentials"
        headers = ["App", "Username", "Password", "TOTP Secret"]
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=i, value=header)
            cell.font = Font(bold=True)
        j = 2
        for app, cred_list in creds.items():
            for cred in cred_list:
                username = cred.get('username', '')
                password = crypto.decrypt_aes256(cred.get('password', ''))
                secret = cred.get('secretco', '')
                if secret:
                    secret = crypto.decrypt_aes256(secret)
                ws.cell(row=j, column=1, value=app)
                ws.cell(row=j, column=2, value=username)
                ws.cell(row=j, column=3, value=password)
                ws.cell(row=j, column=4, value=secret)
                j += 1
        desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop', 'credentials.xlsx')
        wb.save(desktop_path)
        return True, f"File saved to Desktop. Be sure to secure it."
    except Exception as e:
        return False, f"Failed to download Excel"